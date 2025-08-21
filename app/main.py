import asyncio, os, uuid, json, csv
from typing import Optional
from fastapi import FastAPI, UploadFile, File as UploadFileType, HTTPException, Depends, Header, Response
from fastapi.responses import StreamingResponse
from fastapi.middleware.cors import CORSMiddleware
from sqlalchemy.orm import Session
from sqlalchemy import select, func
import openpyxl

from .db import SessionLocal, init_db, File as DBFile, ParsedRow, FileStatus, User
from .auth import create_access_token, decode_token, verify_password, get_password_hash, gen_uuid

init_db()
os.makedirs("uploads", exist_ok=True)

CHUNK_SIZE = int(os.getenv("CHUNK_SIZE", "1048576"))
DEFAULT_PAGE_SIZE = int(os.getenv("PAGE_SIZE", "100"))

app = FastAPI(title="File Parser CRUD API Full")
app.add_middleware(CORSMiddleware, allow_origins=["*"], allow_credentials=True, allow_methods=["*"], allow_headers=["*"])

# Simple dependency to get DB session
def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()

# JWT auth dependency
def get_current_user(authorization: Optional[str] = Header(None)):
    if not authorization:
        raise HTTPException(status_code=401, detail="Missing Authorization header")
    scheme, _, token = authorization.partition(" ")
    if scheme.lower() != "bearer" or not token:
        raise HTTPException(status_code=401, detail="Invalid auth scheme")
    payload = decode_token(token)
    if not payload or "sub" not in payload:
        raise HTTPException(status_code=401, detail="Invalid token")
    db = next(get_db())
    user = db.get(User, payload["sub"])
    if not user:
        raise HTTPException(status_code=401, detail="User not found")
    return user

# SSE subscribers
progress_subscribers = {}

async def notify_progress(file_id: str, payload: dict):
    queues = progress_subscribers.get(file_id, set())
    for q in list(queues):
        await q.put(payload)

@app.post("/auth/signup", status_code=201)
def signup(email: str, password: str):
    db: Session = next(get_db())
    existing = db.execute(select(User).where(User.email == email)).scalar_one_or_none()
    if existing:
        raise HTTPException(status_code=400, detail="Email already registered")
    user = User(id=gen_uuid(), email=email, hashed_password=get_password_hash(password))
    db.add(user); db.commit()
    return {"msg": "created", "user_id": user.id}

@app.post("/auth/login")
def login(email: str, password: str):
    db: Session = next(get_db())
    user = db.execute(select(User).where(User.email == email)).scalar_one_or_none()
    if not user or not verify_password(password, user.hashed_password):
        raise HTTPException(status_code=401, detail="Invalid credentials")
    token = create_access_token({"sub": user.id, "email": user.email})
    return {"access_token": token, "token_type": "bearer"}

@app.post("/files", status_code=201)
async def upload_file(file: UploadFileType = UploadFile(...), current=Depends(get_current_user)):
    file_id = str(uuid.uuid4())
    storage_path = os.path.join("uploads", f"{file_id}_{file.filename}")
    size_bytes = 0

    db: Session = next(get_db())
    db_file = DBFile(id=file_id, filename=file.filename, mime_type=file.content_type, storage_path=storage_path, status=FileStatus.uploading, upload_progress=0, parse_progress=0)
    db.add(db_file); db.commit()

    # Stream upload to disk in chunks and update progress (simulated if content-length unknown)
    with open(storage_path, "wb") as out:
        while True:
            chunk = await file.read(CHUNK_SIZE)
            if not chunk:
                break
            out.write(chunk)
            size_bytes += len(chunk)
            db_file.upload_progress = min(95, (db_file.upload_progress or 0) + 5)
            db.commit()
            await notify_progress(file_id, {"stage": "uploading", "progress": db_file.upload_progress})

    db_file.size_bytes = size_bytes
    db_file.upload_progress = 100
    db_file.status = FileStatus.processing
    db.commit()
    await notify_progress(file_id, {"stage": "uploading", "progress": 100})

    asyncio.create_task(parse_file_async(file_id))
    return {"file_id": file_id, "status": db_file.status.value}

async def parse_file_async(file_id: str):
    db: Session = next(get_db())
    db_file: Optional[DBFile] = db.get(DBFile, file_id)
    if not db_file:
        return
    path = db_file.storage_path
    processed = 0
    try:
        ext = os.path.splitext(path)[1].lower()
        if ext in [".xlsx", ".xls"]:
            # stream using openpyxl to avoid loading whole workbook
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            sheet = wb.active
            total = sheet.max_row if sheet.max_row else 0
            row_index = 0
            for row in sheet.iter_rows(values_only=True):
                if row_index == 0 and any(cell is not None for cell in row):
                    # assume header exists
                    headers = [str(c) if c is not None else f"col{idx}" for idx,c in enumerate(row)]
                    row_index += 1
                    continue
                # map to headers if available
                data = {}
                for i,cell in enumerate(row):
                    key = headers[i] if i < len(headers) else f"col{i}"
                    data[key] = "" if cell is None else str(cell)
                db.add(ParsedRow(file_id=file_id, row_index=row_index, data_json=json.dumps(data)))
                row_index += 1
                processed += 1
                if processed % 50 == 0:
                    db.commit()
                    # update progress estimate
                    pct = int((processed / max(1, total)) * 100) if total>0 else min(99, (db_file.parse_progress or 0)+5)
                    db_file.parse_progress = min(99, pct)
                    db.commit()
                    await notify_progress(file_id, {"stage":"processing","progress":db_file.parse_progress})
            db.commit()
        else:
            # CSV streaming
            total = 0
            try:
                with open(path, "r", encoding="utf-8", errors="ignore") as f:
                    for _ in f:
                        total += 1
            except Exception:
                total = 0
            with open(path, "r", encoding="utf-8", errors="ignore") as f:
                reader = csv.reader(f)
                headers = None
                row_index = 0
                for row in reader:
                    if row_index == 0:
                        headers = [h if h is not None else f"col{idx}" for idx,h in enumerate(row)]
                        row_index += 1
                        continue
                    data = { (headers[i] if i < len(headers) else f"col{i}"): (row[i] if i < len(row) else "") for i in range(len(row)) }
                    db.add(ParsedRow(file_id=file_id, row_index=row_index, data_json=json.dumps(data)))
                    row_index += 1
                    processed += 1
                    if processed % 100 == 0:
                        db.commit()
                        pct = int((processed / max(1, total)) * 100) if total>0 else min(99, (db_file.parse_progress or 0)+5)
                        db_file.parse_progress = min(99, pct)
                        db.commit()
                        await notify_progress(file_id, {"stage":"processing","progress":db_file.parse_progress})
            db.commit()

        db_file.status = FileStatus.ready
        db_file.parse_progress = 100
        db.commit()
        await notify_progress(file_id, {"stage":"processing","progress":100})

    except Exception as e:
        db_file.status = FileStatus.failed
        db_file.error_message = str(e)
        db_file.parse_progress = 0
        db.commit()
        await notify_progress(file_id, {"stage":"failed","error":str(e)})

@app.get("/files/{file_id}/progress")
def get_progress(file_id: str):
    db: Session = next(get_db())
    db_file = db.get(DBFile, file_id)
    if not db_file:
        raise HTTPException(status_code=404, detail="File not found")
    if db_file.status == FileStatus.uploading:
        progress = db_file.upload_progress
    elif db_file.status == FileStatus.processing:
        progress = db_file.parse_progress
    elif db_file.status == FileStatus.ready:
        progress = 100
    else:
        progress = db_file.parse_progress or db_file.upload_progress or 0
    return {"file_id": file_id, "status": db_file.status.value, "progress": progress}

@app.get("/files/{file_id}")
def get_parsed_content(file_id: str, page: int = 1, limit: int = DEFAULT_PAGE_SIZE):
    db: Session = next(get_db())
    db_file = db.get(DBFile, file_id)
    if not db_file:
        raise HTTPException(status_code=404, detail="File not found")
    if db_file.status != FileStatus.ready:
        return Response(content=json.dumps({"message":"File upload or processing in progress. Please try again later."}), status_code=202, media_type="application/json")
    page = max(1, page)
    limit = max(1, min(1000, limit))
    offset = (page - 1) * limit
    total = db.execute(select(func.count()).select_from(ParsedRow).where(ParsedRow.file_id == file_id)).scalar_one()
    rows = db.execute(select(ParsedRow).where(ParsedRow.file_id == file_id).order_by(ParsedRow.row_index).offset(offset).limit(limit)).scalars().all()
    parsed = [json.loads(r.data_json) for r in rows]
    return {"file_id": file_id, "status": db_file.status.value, "page": page, "limit": limit, "total_rows": total, "rows": parsed}

@app.get("/files")
def list_files():
    db: Session = next(get_db())
    items = db.execute(select(DBFile).order_by(DBFile.created_at.desc())).scalars().all()
    return [{"id": f.id, "filename": f.filename, "status": f.status.value, "created_at": f.created_at.isoformat() if f.created_at else None} for f in items]

@app.delete("/files/{file_id}", status_code=204)
def delete_file(file_id: str, current=Depends(get_current_user)):
    db: Session = next(get_db())
    db_file = db.get(DBFile, file_id)
    if not db_file:
        raise HTTPException(status_code=404, detail="File not found")
    try:
        if os.path.exists(db_file.storage_path):
            os.remove(db_file.storage_path)
    except Exception:
        pass
    db.delete(db_file)
    db.commit()
    return Response(status_code=204)

@app.get("/files/{file_id}/events")
async def sse_events(file_id: str):
    q = asyncio.Queue()
    progress_subscribers.setdefault(file_id, set()).add(q)
    async def event_generator():
        try:
            while True:
                msg = await q.get()
                yield f"event: progress\ndata: {json.dumps(msg)}\n\n"
        except asyncio.CancelledError:
            pass
        finally:
            progress_subscribers[file_id].discard(q)
    return StreamingResponse(event_generator(), media_type="text/event-stream")
